"""
ClosePilot Enterprise Demo Pack Generator
Generates: ClosePilot_Enterprise_Demo_Pack.xlsx
Company:   Northstar Manufacturing Ltd
"""

import random
import datetime
import numpy as np
import pandas as pd
from faker import Faker
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers as num_formats
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo

fake = Faker("en_GB")
random.seed(42)
np.random.seed(42)

# ── Company constants ──────────────────────────────────────────────────────────

COMPANY       = "Northstar Manufacturing Ltd"
YEAR_END      = "31 December 2025"
REVENUE       = 8_500_000
COGS          = 7_050_000
GROSS_PROFIT  = REVENUE - COGS           # £1,450,000  (17%)
EBITDA        = 180_000
DEPRECIATION  = 120_000
EBIT          = EBITDA - DEPRECIATION    # £60,000
INTEREST      = 120_000
PBT           = EBIT - INTEREST          # -£60,000 (small loss)
NET_PROFIT    = PBT                      # no tax on loss

EMPLOYEES     = 75
MONTHLY_PAY   = 215_000
ANNUAL_PAY    = MONTHLY_PAY * 12        # £2,580,000

# Balance sheet totals
FA_COST       = 4_100_000
INVENTORY     = 600_000
TRADE_DEBTORS = 850_000
BANK          = 340_000
PREPAYMENTS   = 80_000
OTHER_CA      = 20_000
TOTAL_ASSETS  = FA_COST + INVENTORY + TRADE_DEBTORS + BANK + PREPAYMENTS + OTHER_CA

TRADE_CRED    = 500_000
VAT_CTRL      = 42_000
PAYE          = 35_000
BANK_LOAN     = 3_200_000
ACCRUALS      = 150_000
TOTAL_LIAB    = TRADE_CRED + VAT_CTRL + PAYE + BANK_LOAN + ACCRUALS

SHARE_CAP     = 100_000
RETAINED_EARN = TOTAL_ASSETS - TOTAL_LIAB - SHARE_CAP   # balancing figure
TOTAL_EQUITY  = SHARE_CAP + RETAINED_EARN

# AR deliberate mismatch
AR_AGING_TOTAL  = 895_000   # £45k higher than TB
AR_MISMATCH     = AR_AGING_TOTAL - TRADE_DEBTORS  # £45,000

# VAT mismatch
VAT_BOX1        = 50_400
VAT_BOX4        = 18_700
VAT_BOX6        = 2_850_000
VAT_BOX7        = 1_920_000
VAT_MISMATCH    = VAT_BOX1 - VAT_BOX4 - VAT_CTRL   # £-10,300 difference

# Bank mismatch
BANK_STMT       = 325_000
BANK_UNRECON    = BANK - BANK_STMT       # £15,000

# ── Style helpers ─────────────────────────────────────────────────────────────

DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "BDD7EE"
AMBER      = "FFC000"
RED_FILL   = "FF0000"
GREEN_FILL = "70AD47"
LIGHT_GREY = "F2F2F2"
WHITE      = "FFFFFF"

def hdr(ws, row, col, text, bold=True, bg=MID_BLUE, fg=WHITE, size=11, wrap=False, align="left"):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=bold, color=fg, size=size, name="Calibri")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    return cell

def val(ws, row, col, value, fmt=None, bold=False, bg=None, align="right"):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, name="Calibri", size=10)
    if fmt:
        cell.number_format = fmt
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    return cell

def freeze(ws, row, col):
    ws.freeze_panes = ws.cell(row=row, column=col)

def col_w(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def thin_border():
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def add_table(ws, ref, name, style="TableStyleMedium2"):
    tbl = Table(displayName=name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(name=style, showFirstColumn=False,
        showLastColumn=False, showRowStripes=True)
    ws.add_table(tbl)

GBP = '£#,##0.00'
GBP0 = '£#,##0'
PCT = '0.0%'
INT0 = '#,##0'
DATE_FMT = 'DD/MM/YYYY'

# ── 1. Company Profile ────────────────────────────────────────────────────────

def sheet_company_profile(wb):
    ws = wb.create_sheet("Company Profile")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 40

    hdr(ws, 1, 1, "NORTHSTAR MANUFACTURING LTD", bg=DARK_BLUE, size=14, bold=True)
    ws.merge_cells("A1:B1")
    hdr(ws, 2, 1, "ClosePilot Enterprise Demo Pack", bg=MID_BLUE, size=11)
    ws.merge_cells("A2:B2")

    rows = [
        ("Company Name",       COMPANY),
        ("Company Number",     "03847291"),
        ("Industry",           "Manufacturing — Industrial Components"),
        ("Country",            "United Kingdom"),
        ("Year End",           YEAR_END),
        ("Accounting System",  "Sage 200 Enterprise"),
        ("Currency",           "GBP"),
        ("Employees",          75),
        ("Directors",          2),
        ("Auditors",           "Grant Thornton LLP"),
        ("Bankers",            "Barclays Business Banking"),
        ("VAT Registration",   "GB 123 4567 89"),
        ("Corporation Tax Ref","UTR 12345 67890"),
        ("", ""),
        ("FINANCIAL SUMMARY",  ""),
        ("Annual Revenue",     REVENUE),
        ("Gross Margin",       f"{GROSS_PROFIT/REVENUE:.1%}"),
        ("EBITDA",             EBITDA),
        ("Net Profit / (Loss)",NET_PROFIT),
        ("Total Assets",       TOTAL_ASSETS),
        ("Net Assets",         TOTAL_EQUITY),
        ("Bank Loans",         BANK_LOAN),
        ("", ""),
        ("DEMO PACK PURPOSE",  ""),
        ("Generated by",       "ClosePilot Assurance Platform"),
        ("Purpose",            "Enterprise demonstration — contains deliberate data issues"),
        ("Pack Version",       "v1.0"),
        ("Generated",          datetime.date.today().strftime("%d %B %Y")),
    ]

    for i, (label, value) in enumerate(rows, start=4):
        if label in ("FINANCIAL SUMMARY", "DEMO PACK PURPOSE"):
            hdr(ws, i, 1, label, bg=LIGHT_BLUE, fg=DARK_BLUE, bold=True)
            ws.merge_cells(f"A{i}:B{i}")
            continue
        if label == "":
            continue
        c = ws.cell(row=i, column=1, value=label)
        c.font = Font(bold=True, name="Calibri", size=10)
        c.fill = PatternFill("solid", fgColor=LIGHT_GREY)
        c2 = ws.cell(row=i, column=2, value=value)
        c2.font = Font(name="Calibri", size=10)
        if isinstance(value, int) and value > 1000:
            c2.number_format = GBP0

# ── 2. Trial Balance ──────────────────────────────────────────────────────────

def sheet_trial_balance(wb):
    ws = wb.create_sheet("Trial Balance")

    accounts = [
        # FIXED ASSETS
        ("1000", "Freehold Land & Buildings",              1_800_000, "D"),
        ("1001", "Plant & Machinery – Cost",               3_500_000, "D"),
        ("1002", "Motor Vehicles – Cost",                    420_000, "D"),
        ("1003", "IT Equipment – Cost",                      180_000, "D"),
        ("1004", "Office Fixtures & Fittings",               120_000, "D"),
        ("1010", "Accumulated Depreciation – Buildings",    -120_000, "C"),
        ("1011", "Accumulated Depreciation – P&M",          -900_000, "C"),
        ("1012", "Accumulated Depreciation – Vehicles",     -250_000, "C"),
        ("1013", "Accumulated Depreciation – IT",           -130_000, "C"),
        ("1014", "Accumulated Depreciation – Fixtures",      -50_000, "C"),
        ("1020", "Capital WIP",                               80_000, "D"),
        ("1030", "Goodwill at Cost",                         400_000, "D"),
        ("1031", "Accumulated Amortisation – Goodwill",      -80_000, "C"),
        # CURRENT ASSETS
        ("1100", "Raw Materials Inventory",                  320_000, "D"),
        ("1101", "WIP Inventory",                            180_000, "D"),
        ("1102", "Finished Goods Inventory",                 100_000, "D"),
        ("1110", "Trade Debtors Control",           TRADE_DEBTORS, "D"),  # INJECTED
        ("1111", "Provision for Bad Debts",                  -25_000, "C"),
        ("1112", "Other Debtors",                             18_000, "D"),
        ("1113", "Accrued Income",                            32_000, "D"),
        ("1120", "Prepayments",                     PREPAYMENTS, "D"),
        ("1130", "VAT Receivable",                             8_500, "D"),
        ("1140", "Corporation Tax Asset",                     12_000, "D"),
        ("1150", "Bank – Current Account",              BANK, "D"),      # INJECTED
        ("1151", "Bank – Deposit Account",                    28_000, "D"),
        ("1152", "Petty Cash",                                 1_200, "D"),
        ("1160", "Inter-company Receivable",                  45_000, "D"),
        # CURRENT LIABILITIES
        ("2000", "Trade Creditors Control",                 -TRADE_CRED, "C"),
        ("2001", "Accruals",                               -ACCRUALS, "C"),
        ("2002", "Deferred Income",                          -35_000, "C"),
        ("2010", "VAT Control",                            -VAT_CTRL, "C"),  # INJECTED
        ("2011", "PAYE & NIC Payable",                        -PAYE, "C"),
        ("2012", "Pension Payable",                           -18_000, "C"),
        ("2013", "Corporation Tax Payable",                       0, "C"),
        ("2014", "Dividend Payable",                              0, "C"),
        ("2020", "Bank Overdraft",                                0, "C"),
        ("2030", "HP & Finance Lease – Current",              -24_000, "C"),
        ("2031", "Invoice Discounting",                      -180_000, "C"),
        ("2040", "Other Creditors",                           -14_000, "C"),
        # NON-CURRENT LIABILITIES
        ("2500", "Bank Loan – Barclays",               -BANK_LOAN, "C"),
        ("2501", "HP & Finance Lease – Long Term",         -72_000, "C"),
        ("2510", "Deferred Tax Liability",                   -28_000, "C"),
        # EQUITY
        ("3000", "Called Up Share Capital",             -SHARE_CAP, "C"),
        ("3001", "Share Premium Account",                        0, "C"),
        ("3010", "Revaluation Reserve",                  -120_000, "C"),
        ("3020", "Retained Earnings – B/F",       -RETAINED_EARN - NET_PROFIT, "C"),
        # REVENUE
        ("4000", "Product Sales – UK",                -6_200_000, "C"),
        ("4001", "Product Sales – Export",             -1_800_000, "C"),
        ("4002", "Service Revenue",                      -300_000, "C"),
        ("4003", "Tooling Revenue",                      -180_000, "C"),
        ("4004", "Other Revenue",                         -20_000, "C"),
        ("4010", "Rebates & Discounts",                   180_000, "D"),
        # COST OF SALES
        ("5000", "Raw Materials – Steel",               2_800_000, "D"),
        ("5001", "Raw Materials – Aluminium",             650_000, "D"),
        ("5002", "Raw Materials – Plastics",              420_000, "D"),
        ("5003", "Raw Materials – Other",                 310_000, "D"),
        ("5010", "Direct Labour – Production",          1_350_000, "D"),
        ("5011", "Direct Labour – Assembly",              420_000, "D"),
        ("5012", "Agency & Temporary Labour",             180_000, "D"),
        ("5020", "Production Overhead – Fixed",           280_000, "D"),
        ("5021", "Production Overhead – Variable",        220_000, "D"),
        ("5022", "Factory Rent & Rates",                  145_000, "D"),
        ("5023", "Energy – Production",                    96_000, "D"),
        ("5024", "Machine Maintenance",                    68_000, "D"),
        ("5025", "Subcontract Costs",                     111_000, "D"),
        # GROSS PROFIT CHECK: COGS = 7,050,000
        # OPERATING EXPENSES
        ("6000", "Directors' Salaries",                   180_000, "D"),
        ("6001", "Management Salaries",                   280_000, "D"),
        ("6002", "Sales & Marketing Salaries",            220_000, "D"),
        ("6003", "Admin & Finance Salaries",              150_000, "D"),
        ("6004", "Employer NIC",                           82_000, "D"),
        ("6005", "Employer Pension",                       48_000, "D"),
        ("6010", "Office Rent & Rates",                    72_000, "D"),
        ("6011", "Office Utilities",                       18_000, "D"),
        ("6012", "Telephone & Broadband",                  14_400, "D"),
        ("6013", "IT & Software",                          38_000, "D"),
        ("6014", "Insurance",                              42_000, "D"),
        ("6015", "Legal & Professional",                   56_000, "D"),
        ("6016", "Audit Fees",                             28_000, "D"),
        ("6017", "Marketing & Advertising",                32_000, "D"),
        ("6018", "Travel & Entertaining",                  24_000, "D"),
        ("6019", "Vehicle Running Costs",                  36_000, "D"),
        ("6020", "Sundry Expenses",                        12_600, "D"),
        ("6030", "Depreciation – Buildings",               18_000, "D"),
        ("6031", "Depreciation – P&M",                     72_000, "D"),
        ("6032", "Depreciation – Vehicles",                18_000, "D"),
        ("6033", "Depreciation – IT",                      12_000, "D"),
        # NOTE: £750,000 of plant has zero depreciation injected separately
        ("6040", "Bad Debt Write-off",                      8_400, "D"),
        ("6041", "Bad Debt Provision Movement",              6_000, "D"),
        # FINANCE COSTS
        ("7000", "Bank Interest",                          96_000, "D"),
        ("7001", "HP / Lease Finance Charge",              18_000, "D"),
        ("7002", "Invoice Discounting Charges",             6_000, "D"),
        # CORPORATION TAX
        ("8000", "Corporation Tax Charge",                      0, "D"),
    ]

    # Generate 170 more minor accounts to reach 250+
    extra_accounts = []
    prefixes = [
        ("5030", "Packaging Materials",          28_000),
        ("5031", "Quality & Testing",            14_400),
        ("5032", "H&S & Compliance",             12_000),
        ("5033", "Waste Disposal",                9_600),
        ("5034", "Calibration Costs",             6_000),
        ("5035", "Stock Provision",              12_000),
        ("5036", "Scrap Recovery",               -8_400),
        ("6050", "Training & Development",        8_400),
        ("6051", "Recruitment Costs",             6_000),
        ("6052", "Employee Welfare",              3_600),
        ("6053", "Workwear & PPE",                4_800),
        ("6054", "Cleaning & Hygiene",            6_000),
        ("6055", "Security",                      9_600),
        ("6056", "Water Rates",                   7_200),
        ("6057", "Postage & Courier",             4_200),
        ("6058", "Stationery & Office",           3_600),
        ("6059", "Subscriptions",                 4_800),
        ("6060", "Bank Charges",                  3_600),
        ("6061", "Credit Card Charges",           1_800),
        ("6062", "Foreign Exchange Loss",         2_400),
        ("6063", "Penalty & Interest",            1_200),
        ("6064", "Donations",                     2_400),
        ("6065", "Staff Entertainment",           3_600),
        ("6066", "Client Entertainment",          4_800),
        ("6067", "Conference & Events",           6_000),
        ("6068", "Research & Development",       24_000),
        ("6069", "Patent & IP Costs",             4_800),
        ("6070", "Regulatory & Compliance",       7_200),
        ("6071", "Environmental Levy",            3_600),
        ("6072", "Carbon Credits",                2_400),
    ]
    for code, name, bal in prefixes:
        extra_accounts.append((code, name, bal, "D" if bal > 0 else "C"))

    # Add 120 minor balance sheet accounts (all near-zero)
    asset_names  = ["Deposit - Landlord","Security Bond","Loan – Employee","Advance – Supplier",
                     "Sundry Debtors","Accrued Interest Receivable","Withholding Tax Asset",
                     "Deferred Financing Cost","RDEC Receivable","Grant Receivable"]
    liab_names   = ["Customer Deposits","Deferred Grant","Warranty Provision","Legal Provision",
                     "Restructuring Reserve","Onerous Lease Provision","Holiday Pay Accrual",
                     "Bonus Accrual","Commission Payable","HMRC Other"]

    for i, name in enumerate(asset_names * 5, start=0):
        code = f"1{900 + i:03d}"
        bal  = random.randint(500, 15_000)
        extra_accounts.append((code, f"{name} {i+1:02d}", bal, "D"))

    for i, name in enumerate(liab_names * 5, start=0):
        code = f"2{600 + i:03d}"
        bal  = random.randint(500, 12_000)
        extra_accounts.append((code, f"{name} {i+1:02d}", -bal, "C"))

    all_accounts = accounts + extra_accounts

    # Force balance: compute running total, add a balancing entry to retained earnings
    total = sum(a[2] for a in all_accounts)
    if abs(total) > 1:
        all_accounts.append(("3021", "Retained Earnings – Current Year Adj", -total, "C" if total > 0 else "D"))

    # Write sheet
    ws = wb.create_sheet("Trial Balance")
    ws.sheet_view.showGridLines = False
    freeze(ws, 4, 1)

    # Title
    ws.merge_cells("A1:E1")
    hdr(ws, 1, 1, f"{COMPANY}  |  Trial Balance  |  Year ended {YEAR_END}", bg=DARK_BLUE, size=12)
    ws.row_dimensions[1].height = 24

    # Column headers
    headers = ["Account Code", "Account Name", "Balance (£)", "Dr/Cr", "Category"]
    for ci, h in enumerate(headers, 1):
        hdr(ws, 3, ci, h, bg=MID_BLUE)

    def get_category(code):
        c = int(code[:2]) if code[:2].isdigit() else 99
        if c <= 19: return "Fixed Assets"
        if c <= 16: return "Current Assets"
        if c <= 20: return "Current Liabilities"
        if c <= 25: return "Non-Current Liabilities"
        if c <= 30: return "Equity"
        if c <= 40: return "Revenue"
        if c <= 59: return "Cost of Sales"
        if c <= 69: return "Operating Expenses"
        if c <= 79: return "Finance Costs"
        return "Taxation"

    cat_colors = {
        "Fixed Assets": "DEEAF1", "Current Assets": "E2EFDA", "Current Liabilities": "FCE4D6",
        "Non-Current Liabilities": "FFDFD6", "Equity": "EBE4F0", "Revenue": "E2EFDA",
        "Cost of Sales": "FFF2CC", "Operating Expenses": "FFF2CC", "Finance Costs": "FCE4D6",
        "Taxation": "F2F2F2",
    }

    dr_total = cr_total = 0
    for ri, (code, name, balance, drcr) in enumerate(all_accounts, start=4):
        abs_bal = abs(balance)
        cat = get_category(code)
        bg_col = cat_colors.get(cat, WHITE)
        ws.cell(ri, 1, code).font = Font(name="Calibri", size=10)
        ws.cell(ri, 2, name).font = Font(name="Calibri", size=10)
        c = ws.cell(ri, 3, abs_bal)
        c.number_format = GBP0
        c.font = Font(name="Calibri", size=10)
        c.alignment = Alignment(horizontal="right")
        ws.cell(ri, 4, drcr).font = Font(name="Calibri", size=10, bold=True,
            color="1F497D" if drcr == "D" else "C00000")
        ws.cell(ri, 5, cat).font = Font(name="Calibri", size=10)
        ws.cell(ri, 5).fill = PatternFill("solid", fgColor=bg_col)
        if drcr == "D": dr_total += abs_bal
        else:           cr_total += abs_bal

    # Totals
    tot_row = 4 + len(all_accounts)
    hdr(ws, tot_row, 1, "TOTALS", bg=DARK_BLUE)
    hdr(ws, tot_row, 2, f"Accounts: {len(all_accounts)}", bg=DARK_BLUE)
    c_dr = ws.cell(tot_row, 3, dr_total)
    c_dr.number_format = GBP0; c_dr.font = Font(bold=True, name="Calibri", size=10, color=WHITE)
    c_dr.fill = PatternFill("solid", fgColor=DARK_BLUE)
    c_dr.alignment = Alignment(horizontal="right")

    diff = dr_total - cr_total
    balance_label = "✓ BALANCED" if abs(diff) < 2 else f"⚠ DIFFERENCE £{diff:,.0f}"
    c_bal = ws.cell(tot_row, 4, balance_label)
    c_bal.font = Font(bold=True, name="Calibri", size=10, color=WHITE)
    c_bal.fill = PatternFill("solid", fgColor=GREEN_FILL if abs(diff) < 2 else "C00000")

    col_w(ws, {"A": 14, "B": 45, "C": 18, "D": 8, "E": 22})

# ── 3. Balance Sheet ──────────────────────────────────────────────────────────

def sheet_balance_sheet(wb):
    ws = wb.create_sheet("Balance Sheet")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:C1")
    hdr(ws, 1, 1, f"{COMPANY}  |  Balance Sheet  |  {YEAR_END}", bg=DARK_BLUE, size=12)

    sections = [
        ("FIXED ASSETS", None, None),
        ("Freehold Land & Buildings – NBV",    1_680_000, None),
        ("Plant & Machinery – NBV",            2_600_000, None),
        ("Motor Vehicles – NBV",                 170_000, None),
        ("IT Equipment – NBV",                    50_000, None),
        ("Office Fixtures – NBV",                 70_000, None),
        ("Capital WIP",                           80_000, None),
        ("Goodwill – NBV",                       320_000, None),
        ("Intangible Assets – NBV",              130_000, None),
        ("TOTAL FIXED ASSETS",                 4_100_000, DARK_BLUE),
        ("", None, None),
        ("CURRENT ASSETS", None, None),
        ("Inventories",                          INVENTORY, None),
        ("Trade Debtors",                   TRADE_DEBTORS, None),
        ("Other Debtors & Accrued Income",        50_000, None),
        ("Prepayments",                         PREPAYMENTS, None),
        ("Bank & Cash",                           BANK, None),
        ("TOTAL CURRENT ASSETS",            INVENTORY + TRADE_DEBTORS + 50_000 + PREPAYMENTS + BANK, MID_BLUE),
        ("", None, None),
        ("TOTAL ASSETS",                    TOTAL_ASSETS, DARK_BLUE),
        ("", None, None),
        ("CURRENT LIABILITIES", None, None),
        ("Trade Creditors",                   -TRADE_CRED, None),
        ("VAT Control",                         -VAT_CTRL, None),
        ("PAYE & NIC Payable",                     -PAYE, None),
        ("Accruals",                            -ACCRUALS, None),
        ("Invoice Discounting",                  -180_000, None),
        ("HP Finance – Current",                  -24_000, None),
        ("Other Creditors",                       -14_000, None),
        ("TOTAL CURRENT LIABILITIES",           -969_000, MID_BLUE),
        ("", None, None),
        ("NON-CURRENT LIABILITIES", None, None),
        ("Bank Loan – Barclays",              -BANK_LOAN, None),
        ("HP Finance – Long Term",                -72_000, None),
        ("Deferred Tax",                          -28_000, None),
        ("TOTAL NON-CURRENT LIABILITIES",      -3_300_000, MID_BLUE),
        ("", None, None),
        ("TOTAL LIABILITIES",                  -TOTAL_LIAB, DARK_BLUE),
        ("", None, None),
        ("NET ASSETS",             TOTAL_ASSETS - TOTAL_LIAB, "1F3864"),
        ("", None, None),
        ("EQUITY", None, None),
        ("Called Up Share Capital",              SHARE_CAP, None),
        ("Share Premium",                              0, None),
        ("Revaluation Reserve",                  120_000, None),
        ("Retained Earnings",                 RETAINED_EARN, None),
        ("TOTAL EQUITY",                       TOTAL_EQUITY, DARK_BLUE),
        ("", None, None),
        ("BALANCE SHEET CHECK",    TOTAL_ASSETS - TOTAL_LIAB - TOTAL_EQUITY, None),
    ]

    hdr(ws, 3, 1, "Account", bg=MID_BLUE)
    hdr(ws, 3, 2, "£", bg=MID_BLUE, align="right")
    hdr(ws, 3, 3, "Notes", bg=MID_BLUE)

    for ri, (label, value, color) in enumerate(sections, start=4):
        if label == "":
            ws.row_dimensions[ri].height = 6; continue
        is_total = label.startswith("TOTAL") or label in ("NET ASSETS", "BALANCE SHEET CHECK","EQUITY","FIXED ASSETS","CURRENT ASSETS","CURRENT LIABILITIES","NON-CURRENT LIABILITIES")
        bg = color if color else (LIGHT_GREY if is_total else WHITE)
        fg = WHITE if color and color != LIGHT_GREY else "000000"
        c1 = ws.cell(ri, 1, label)
        c1.font = Font(bold=is_total, name="Calibri", size=10, color=fg)
        c1.fill = PatternFill("solid", fgColor=bg)
        if value is not None:
            c2 = ws.cell(ri, 2, value)
            c2.number_format = GBP0
            c2.font = Font(bold=is_total, name="Calibri", size=10, color=fg)
            c2.fill = PatternFill("solid", fgColor=bg)
            c2.alignment = Alignment(horizontal="right")
            if label == "BALANCE SHEET CHECK":
                c2.fill = PatternFill("solid", fgColor=GREEN_FILL if abs(value) < 2 else "C00000")
                ws.cell(ri, 3, "✓ Assets = Liabilities + Equity" if abs(value) < 2 else "⚠ DOES NOT BALANCE")
        else:
            c1.fill = PatternFill("solid", fgColor=MID_BLUE if is_total else LIGHT_BLUE)

    col_w(ws, {"A": 42, "B": 18, "C": 36})

# ── 4. Profit & Loss ──────────────────────────────────────────────────────────

def sheet_pnl(wb):
    ws = wb.create_sheet("Profit & Loss")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:C1")
    hdr(ws, 1, 1, f"{COMPANY}  |  Profit & Loss  |  Year ended {YEAR_END}", bg=DARK_BLUE, size=12)
    hdr(ws, 3, 1, "Category", bg=MID_BLUE)
    hdr(ws, 3, 2, "£", bg=MID_BLUE, align="right")
    hdr(ws, 3, 3, "% Revenue", bg=MID_BLUE, align="right")

    rows = [
        ("REVENUE", None, None, DARK_BLUE),
        ("Product Sales – UK",          6_200_000, None, None),
        ("Product Sales – Export",      1_800_000, None, None),
        ("Service Revenue",               300_000, None, None),
        ("Tooling Revenue",               180_000, None, None),
        ("Other Revenue",                  20_000, None, None),
        ("Revenue Rebates",              -180_000, None, None),
        ("NET REVENUE",                 REVENUE, None, MID_BLUE),
        ("", None, None, None),
        ("COST OF SALES", None, None, DARK_BLUE),
        ("Raw Materials",              4_180_000, None, None),
        ("Direct Labour",              1_770_000, None, None),
        ("Agency & Temp Labour",         180_000, None, None),
        ("Production Overhead",          500_000, None, None),
        ("Factory Rent & Rates",         145_000, None, None),
        ("Energy – Production",           96_000, None, None),
        ("Machine Maintenance",           68_000, None, None),
        ("Subcontract Costs",            111_000, None, None),
        ("TOTAL COST OF SALES",        COGS, None, MID_BLUE),
        ("", None, None, None),
        ("GROSS PROFIT",               GROSS_PROFIT, None, "1F3864"),
        ("", None, None, None),
        ("OPERATING EXPENSES", None, None, DARK_BLUE),
        ("Payroll – Overheads",          780_000, None, None),
        ("Employer NIC & Pension",       130_000, None, None),
        ("Office Rent & Rates",           72_000, None, None),
        ("IT & Software",                 38_000, None, None),
        ("Legal & Professional",          84_000, None, None),
        ("Marketing & Advertising",       32_000, None, None),
        ("Insurance",                     42_000, None, None),
        ("Travel & Entertaining",         24_000, None, None),
        ("Vehicle Running Costs",         36_000, None, None),
        ("Utilities – Office",            18_000, None, None),
        ("Bad Debt Provision",            14_400, None, None),
        ("R&D Costs",                     24_000, None, None),
        ("Other Overheads",              (GROSS_PROFIT - EBITDA) - 780_000 - 130_000 - 72_000
                                         - 38_000 - 84_000 - 32_000 - 42_000 - 24_000
                                         - 36_000 - 18_000 - 14_400 - 24_000, None, None),
        ("TOTAL OPERATING EXPENSES",   GROSS_PROFIT - EBITDA, None, MID_BLUE),
        ("", None, None, None),
        ("EBITDA",                      EBITDA, None, "1F3864"),
        ("Depreciation & Amortisation",-DEPRECIATION, None, None),
        ("EBIT",                         EBIT, None, MID_BLUE),
        ("", None, None, None),
        ("FINANCE COSTS", None, None, DARK_BLUE),
        ("Bank Interest",               -96_000, None, None),
        ("HP Finance Charges",          -18_000, None, None),
        ("Invoice Discounting",          -6_000, None, None),
        ("TOTAL FINANCE COSTS",        -INTEREST, None, MID_BLUE),
        ("", None, None, None),
        ("PROFIT / (LOSS) BEFORE TAX", NET_PROFIT, None, DARK_BLUE),
        ("Corporation Tax",                    0, None, None),
        ("NET PROFIT / (LOSS)",        NET_PROFIT, None, "1F3864"),
    ]

    for ri, (label, value, pct, color) in enumerate(rows, start=4):
        if label == "":
            ws.row_dimensions[ri].height = 6; continue
        is_total = any(label.startswith(x) for x in ("TOTAL","GROSS","EBITDA","EBIT","NET","PROFIT"))
        bg = color if color else (LIGHT_GREY if is_total else WHITE)
        fg = WHITE if color and color not in (LIGHT_GREY, WHITE) else "000000"
        c1 = ws.cell(ri, 1, label)
        c1.font = Font(bold=is_total, name="Calibri", size=10, color=fg)
        c1.fill = PatternFill("solid", fgColor=bg)
        if value is not None:
            c2 = ws.cell(ri, 2, value)
            c2.number_format = GBP0
            c2.font = Font(bold=is_total, name="Calibri", size=10, color=fg)
            c2.fill = PatternFill("solid", fgColor=bg)
            c2.alignment = Alignment(horizontal="right")
            c3 = ws.cell(ri, 3, value / REVENUE)
            c3.number_format = PCT
            c3.font = Font(name="Calibri", size=10, color=fg)
            c3.fill = PatternFill("solid", fgColor=bg)
            c3.alignment = Alignment(horizontal="right")
        else:
            c1.fill = PatternFill("solid", fgColor=MID_BLUE if is_total else LIGHT_BLUE)

    col_w(ws, {"A": 42, "B": 18, "C": 14})

# ── 5. 12-Month P&L History ───────────────────────────────────────────────────

def sheet_12month(wb):
    ws = wb.create_sheet("12 Month P&L History")
    months = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

    # Seasonality index (manufacturing – Q4 strongest)
    season = [0.072,0.070,0.082,0.083,0.086,0.086,0.075,0.088,0.090,0.092,0.088,0.088]
    revenue_m  = [REVENUE * s for s in season]
    cogs_m     = [r * 0.83 for r in revenue_m]
    gp_m       = [r - c for r, c in zip(revenue_m, cogs_m)]
    payroll_m  = [ANNUAL_PAY / 12 + random.randint(-5000, 5000) for _ in months]
    rent_m     = [72_000 / 12] * 12
    utilities_m= [18_000 / 12 + random.randint(-500, 2000) for _ in months]
    ebitda_m   = [gp - pay - ren - util - (GROSS_PROFIT - EBITDA - ANNUAL_PAY - 72_000 - 18_000) / 12
                  for gp, pay, ren, util in zip(gp_m, payroll_m, rent_m, utilities_m)]

    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:N1")
    hdr(ws, 1, 1, f"{COMPANY}  |  12 Month P&L History  |  Jan–Dec 2025", bg=DARK_BLUE, size=12)

    row_labels = ["Revenue","COGS","Gross Profit","Gross Margin %","Payroll","Rent","Utilities","EBITDA","EBITDA Margin %"]
    hdr(ws, 3, 1, "Metric", bg=MID_BLUE)
    for mi, m in enumerate(months, 2):
        hdr(ws, 3, mi, m, bg=MID_BLUE, align="right")
    hdr(ws, 3, 14, "FY Total", bg=DARK_BLUE, align="right")

    data_rows = [revenue_m, cogs_m, gp_m,
                 [g/r for g,r in zip(gp_m, revenue_m)],
                 payroll_m, rent_m, utilities_m, ebitda_m,
                 [e/r for e,r in zip(ebitda_m, revenue_m)]]

    is_pct = [False,False,False,True,False,False,False,False,True]

    for li, (label, drow, pct) in enumerate(zip(row_labels, data_rows, is_pct), start=4):
        is_total = label in ("Gross Profit","EBITDA")
        bg = MID_BLUE if is_total else (LIGHT_GREY if li % 2 == 0 else WHITE)
        fg = WHITE if is_total else "000000"
        c = ws.cell(li, 1, label)
        c.font = Font(bold=is_total, name="Calibri", size=10, color=fg)
        c.fill = PatternFill("solid", fgColor=bg)
        for mi, v in enumerate(drow, 2):
            cell = ws.cell(li, mi, v)
            cell.number_format = PCT if pct else GBP0
            cell.font = Font(bold=is_total, name="Calibri", size=10, color=fg)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="right")
        fy = sum(drow) if not pct else sum(drow)/12
        tc = ws.cell(li, 14, fy)
        tc.number_format = PCT if pct else GBP0
        tc.font = Font(bold=True, name="Calibri", size=10, color=WHITE)
        tc.fill = PatternFill("solid", fgColor=DARK_BLUE)
        tc.alignment = Alignment(horizontal="right")

    for ci in range(1, 15):
        ws.column_dimensions[get_column_letter(ci)].width = 14 if ci > 1 else 22

# ── 6. Budget vs Actual ───────────────────────────────────────────────────────

def sheet_budget(wb):
    ws = wb.create_sheet("Budget vs Actual")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    hdr(ws, 1, 1, f"{COMPANY}  |  Budget vs Actual  |  Year ended {YEAR_END}", bg=DARK_BLUE, size=12)

    headers = ["Category", "Budget £", "Actual £", "Variance £", "Variance %", "RAG"]
    for ci, h in enumerate(headers, 1):
        hdr(ws, 3, ci, h, bg=MID_BLUE, align="right" if ci > 1 else "left")

    items = [
        ("Revenue",                 8_900_000, REVENUE),
        ("Cost of Sales",          (7_200_000),(COGS)),
        ("Gross Profit",            1_700_000, GROSS_PROFIT),
        ("Payroll – Overheads",      (760_000),(780_000)),
        ("Marketing",               (30_000),  (32_000)),
        ("IT & Software",           (35_000),  (38_000)),
        ("Legal & Professional",    (75_000),  (84_000)),
        ("Rent & Rates",            (72_000),  (72_000)),
        ("Insurance",               (40_000),  (42_000)),
        ("Other Overheads",        (388_000), (GROSS_PROFIT - EBITDA - 780_000 - 32_000 - 38_000 - 84_000 - 72_000 - 42_000)),
        ("EBITDA",                   300_000,  EBITDA),
        ("Depreciation",           (120_000), (-DEPRECIATION)),
        ("EBIT",                     180_000,   EBIT),
        ("Finance Costs",          (110_000),  (-INTEREST)),
        ("Net Profit / (Loss)",      70_000,   NET_PROFIT),
    ]

    for ri, (cat, budget, actual) in enumerate(items, start=4):
        var = actual - budget
        var_pct = var / abs(budget) if budget else 0
        is_ok = abs(var_pct) < 0.05
        is_total = cat in ("Gross Profit","EBITDA","EBIT","Net Profit / (Loss)")
        bg = MID_BLUE if is_total else WHITE
        fg = WHITE if is_total else "000000"
        rag = "🟢 On Track" if is_ok else ("🟡 Monitor" if abs(var_pct) < 0.15 else "🔴 Off Track")
        ws.cell(ri, 1, cat).font = Font(bold=is_total, name="Calibri", size=10, color=fg)
        ws.cell(ri, 1).fill = PatternFill("solid", fgColor=bg)
        for ci, v in enumerate([budget, actual, var], 2):
            cell = ws.cell(ri, ci, v); cell.number_format = GBP0
            cell.font = Font(bold=is_total, name="Calibri", size=10, color=fg)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.alignment = Alignment(horizontal="right")
        vp = ws.cell(ri, 5, var_pct); vp.number_format = PCT
        vp.font = Font(bold=is_total, name="Calibri", size=10,
            color=("C00000" if var_pct < -0.05 and not is_total else (fg)))
        vp.fill = PatternFill("solid", fgColor=bg); vp.alignment = Alignment(horizontal="right")
        ws.cell(ri, 6, rag).font = Font(name="Calibri", size=10)

    col_w(ws, {"A": 35, "B": 18, "C": 18, "D": 18, "E": 14, "F": 16})

# ── 7. AR Aging ───────────────────────────────────────────────────────────────

def sheet_ar_aging(wb):
    ws = wb.create_sheet("AR Aging")
    ws.sheet_view.showGridLines = False
    freeze(ws, 4, 1)

    ws.merge_cells("A1:J1")
    hdr(ws, 1, 1, f"{COMPANY}  |  AR Aging  |  {YEAR_END}", bg=DARK_BLUE, size=12)

    # Deliberate: three injected customers
    injected = [
        ("Dunlop Retail Ltd",        "CUST001", 245_000, 132, "High",   "Active",  450_000, "Late payments – escalate"),
        ("Northern Build Systems",   "CUST002", 175_000, 95,  "Medium", "Active",  250_000, "Part payment received"),
        ("Tyne Distribution Co",     "CUST003", 145_000, 61,  "Medium", "Active",  200_000, "Invoice disputed"),
    ]

    # Generate 47 more customers
    customers = list(injected)
    remaining = AR_AGING_TOTAL - sum(r[2] for r in injected)  # ~330,000
    for i in range(4, 51):
        balance = random.randint(1_000, max(2_000, remaining // (51 - i)))
        remaining -= balance
        if remaining < 0: remaining = 0
        days = random.choice([15, 21, 30, 45, 60, 75, 90, 120])
        risk = "High" if days >= 90 else ("Medium" if days >= 60 else "Low")
        credit_lim = balance * random.uniform(1.2, 3.0)
        customers.append((
            fake.company()[:40], f"CUST{i:03d}", balance,
            days, risk, "Active", int(credit_lim), random.choice(["","","","Payment overdue","Send statement"])
        ))

    # Ensure last customer fills any remaining gap
    if remaining > 0:
        customers.append((
            "Miscellaneous Debtors", "CUST999", remaining, 30, "Low", "Active", remaining * 2, ""
        ))

    headers = ["Account Code","Customer Name","Outstanding £","Days Overdue","Risk","Status","Credit Limit £","Collector Notes","Overdue Flag","Credit Breach"]
    for ci, h in enumerate(headers, 1):
        hdr(ws, 3, ci, h, bg=MID_BLUE, align="right" if ci > 2 else "left")

    total_bal = 0
    for ri, row in enumerate(customers, start=4):
        acct, name, bal, days, risk, status, clim, note = row[1], row[0], row[2], row[3], row[4], row[5], row[6], row[7]
        total_bal += bal
        breach = "YES" if bal > clim else ""
        overdue = "OVERDUE" if days > 30 else ""
        row_data = [acct, name, bal, days, risk, status, clim, note, overdue, breach]
        for ci, v in enumerate(row_data, 1):
            c = ws.cell(ri, ci, v)
            c.font = Font(name="Calibri", size=10)
            if ci in (3, 7): c.number_format = GBP0; c.alignment = Alignment(horizontal="right")
            if ci == 4: c.alignment = Alignment(horizontal="right")
            if ci == 5:
                c.font = Font(name="Calibri", size=10, bold=True,
                    color="C00000" if risk == "High" else ("FF8C00" if risk == "Medium" else "70AD47"))
            if ri == 4 and name == "Dunlop Retail Ltd":
                c.fill = PatternFill("solid", fgColor="FCE4D6")

    # Total row
    tot = 4 + len(customers)
    hdr(ws, tot, 1, "TOTAL AR AGING", bg=DARK_BLUE)
    c = ws.cell(tot, 3, total_bal); c.number_format = GBP0
    c.font = Font(bold=True, name="Calibri", size=10, color=WHITE)
    c.fill = PatternFill("solid", fgColor=DARK_BLUE); c.alignment = Alignment(horizontal="right")

    # Mismatch note
    hdr(ws, tot + 2, 1, "⚠ RECONCILIATION NOTE", bg=AMBER, fg="000000")
    ws.cell(tot + 2, 3, f"AR Aging Total: £{total_bal:,.0f}")
    ws.cell(tot + 3, 1, "Debtors Control (TB):").font = Font(bold=True, name="Calibri")
    ws.cell(tot + 3, 3, TRADE_DEBTORS).number_format = GBP0
    ws.cell(tot + 4, 1, "DIFFERENCE:").font = Font(bold=True, name="Calibri", color="C00000")
    diff_cell = ws.cell(tot + 4, 3, total_bal - TRADE_DEBTORS)
    diff_cell.number_format = GBP0; diff_cell.font = Font(bold=True, name="Calibri", color="C00000")

    col_w(ws, {"A": 12, "B": 38, "C": 18, "D": 14, "E": 10, "F": 10, "G": 18, "H": 32, "I": 14, "J": 14})
    add_table(ws, f"A3:J{tot - 1}", "ARAgingTable")

# ── 8. AP Aging ───────────────────────────────────────────────────────────────

def sheet_ap_aging(wb):
    ws = wb.create_sheet("AP Aging")
    ws.sheet_view.showGridLines = False
    freeze(ws, 4, 1)

    ws.merge_cells("A1:H1")
    hdr(ws, 1, 1, f"{COMPANY}  |  AP Aging  |  {YEAR_END}", bg=DARK_BLUE, size=12)

    # Injected suppliers with deliberate issues
    injected = [
        ("SUPP001", "SteelCo Ltd",             248_000, 75,  "Medium", "Normal",   "Supplier concentration risk"),
        ("SUPP002", "Power Systems UK",         142_000, 45,  "Low",    "Normal",   ""),
        ("SUPP003", "ABC Services Ltd",          12_400, 30,  "Low",    "Normal",   ""),
        ("SUPP003", "ABC Services Ltd",          12_400, 30,  "Low",    "Duplicate","⚠ DUPLICATE INVOICE — same ref"),
        ("SUPP004", "Mr John Smith",              4_950, 22,  "Low",    "Personal", "⚠ PERSONAL PAYEE — IR35 review"),
        ("SUPP005", "Meridian Tools Ltd",         8_200, 185, "High",   "Overdue",  "⚠ 185 DAYS — consider write-back"),
        ("SUPP006", "Global Freight Co",         34_000, 90,  "High",   "Overdue",  "Payment plan agreed"),
    ]

    # Add 43 more suppliers
    suppliers = list(injected)
    remaining = TRADE_CRED - sum(r[2] for r in injected[:3]) - 4_950 - 8_200 - 34_000
    for i in range(7, 51):
        bal = random.randint(500, max(1000, remaining // (51 - i)))
        remaining -= bal
        if remaining < 0: remaining = 0
        days = random.choice([7, 14, 21, 30, 45, 60])
        risk = "High" if days >= 60 else ("Medium" if days >= 45 else "Low")
        suppliers.append((f"SUPP{i:03d}", fake.company()[:40], bal, days, risk, "Normal", ""))

    headers = ["Account Code","Supplier Name","Outstanding £","Days Aged","Risk","Status","Notes"]
    for ci, h in enumerate(headers, 1):
        hdr(ws, 3, ci, h, bg=MID_BLUE, align="right" if ci > 2 else "left")

    total = 0
    for ri, (acct, name, bal, days, risk, status, note) in enumerate(suppliers, start=4):
        total += bal
        row_data = [acct, name, bal, days, risk, status, note]
        for ci, v in enumerate(row_data, 1):
            c = ws.cell(ri, ci, v)
            c.font = Font(name="Calibri", size=10)
            if ci == 3: c.number_format = GBP0; c.alignment = Alignment(horizontal="right")
            if status in ("Duplicate","Personal"):
                c.fill = PatternFill("solid", fgColor="FCE4D6")
                c.font = Font(name="Calibri", size=10, bold=True)
            elif days >= 90:
                c.fill = PatternFill("solid", fgColor="FFF2CC")

    tot_row = 4 + len(suppliers)
    hdr(ws, tot_row, 1, "TOTAL AP AGING", bg=DARK_BLUE)
    c = ws.cell(tot_row, 3, total); c.number_format = GBP0
    c.font = Font(bold=True, name="Calibri", size=10, color=WHITE)
    c.fill = PatternFill("solid", fgColor=DARK_BLUE); c.alignment = Alignment(horizontal="right")

    col_w(ws, {"A": 12, "B": 40, "C": 18, "D": 12, "E": 10, "F": 12, "G": 45})
    add_table(ws, f"A3:G{tot_row - 1}", "APAgingTable")

# ── 9. VAT Return ─────────────────────────────────────────────────────────────

def sheet_vat_return(wb):
    ws = wb.create_sheet("VAT Return")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:D1")
    hdr(ws, 1, 1, f"{COMPANY}  |  VAT Return  |  Q4 2025 (Oct–Dec)", bg=DARK_BLUE, size=12)

    boxes = [
        ("Box 1", "VAT due on sales and other outputs",      VAT_BOX1, "Output"),
        ("Box 2", "VAT due on acquisitions from EC states",      0,    "Output"),
        ("Box 3", "Total VAT due (Box 1 + Box 2)",           VAT_BOX1, "Output"),
        ("Box 4", "VAT reclaimed on purchases and inputs",   VAT_BOX4, "Input"),
        ("Box 5", "Net VAT to pay / reclaim (Box 3 – Box 4)", VAT_BOX1 - VAT_BOX4, "Net"),
        ("Box 6", "Total value of sales (excl VAT)",         VAT_BOX6, "Sales"),
        ("Box 7", "Total value of purchases (excl VAT)",     VAT_BOX7, "Purchases"),
        ("Box 8", "Total value of EC supplies",                    0,  "EU"),
        ("Box 9", "Total value of EC acquisitions",                0,  "EU"),
    ]

    hdr(ws, 3, 1, "Box", bg=MID_BLUE); hdr(ws, 3, 2, "Description", bg=MID_BLUE)
    hdr(ws, 3, 3, "Amount £", bg=MID_BLUE, align="right"); hdr(ws, 3, 4, "Type", bg=MID_BLUE)

    for ri, (box, desc, amt, btype) in enumerate(boxes, start=4):
        is_net = btype == "Net"
        bg = MID_BLUE if is_net else WHITE
        fg = WHITE if is_net else "000000"
        ws.cell(ri, 1, box).font = Font(bold=True, name="Calibri", size=10, color=fg)
        ws.cell(ri, 1).fill = PatternFill("solid", fgColor=bg)
        ws.cell(ri, 2, desc).font = Font(name="Calibri", size=10, color=fg)
        ws.cell(ri, 2).fill = PatternFill("solid", fgColor=bg)
        c = ws.cell(ri, 3, amt); c.number_format = GBP0
        c.font = Font(bold=is_net, name="Calibri", size=10, color=fg)
        c.fill = PatternFill("solid", fgColor=bg); c.alignment = Alignment(horizontal="right")
        ws.cell(ri, 4, btype).font = Font(name="Calibri", size=10, color=fg)
        ws.cell(ri, 4).fill = PatternFill("solid", fgColor=bg)

    # Reconciliation section
    ws.cell(15, 1, "RECONCILIATION").font = Font(bold=True, name="Calibri", size=11)
    recon = [
        ("VAT Control Account (TB)",              VAT_CTRL),
        ("Box 5 Net VAT (Return)",      VAT_BOX1 - VAT_BOX4),
        ("DIFFERENCE",                  VAT_CTRL - (VAT_BOX1 - VAT_BOX4)),
    ]
    for ri, (label, val_) in enumerate(recon, start=16):
        is_diff = label == "DIFFERENCE"
        ws.cell(ri, 1, label).font = Font(bold=is_diff, name="Calibri", size=10, color="C00000" if is_diff else "000000")
        c = ws.cell(ri, 3, val_); c.number_format = GBP0; c.alignment = Alignment(horizontal="right")
        c.font = Font(bold=is_diff, name="Calibri", size=10, color="C00000" if is_diff else "000000")
        if is_diff:
            ws.cell(ri, 4, "⚠ RECONCILIATION FAILURE — investigate before submission").font = Font(bold=True, color="C00000")

    col_w(ws, {"A": 8, "B": 55, "C": 18, "D": 55})

# ── 10. VAT Transactions ──────────────────────────────────────────────────────

def sheet_vat_transactions(wb):
    ws = wb.create_sheet("VAT Transactions")
    ws.sheet_view.showGridLines = False
    freeze(ws, 4, 1)
    ws.merge_cells("A1:H1")
    hdr(ws, 1, 1, f"{COMPANY}  |  VAT Transactions  |  Q4 2025", bg=DARK_BLUE, size=12)

    headers = ["Date","Supplier / Customer","Net £","VAT £","Gross £","VAT Code","Type","Notes"]
    for ci, h in enumerate(headers, 1):
        hdr(ws, 3, ci, h, bg=MID_BLUE, align="right" if ci > 2 else "left")

    # Deliberately injected transactions
    injected = [
        (datetime.date(2025, 10, 15), "Google Ireland Ltd",      1_800, 0,      1_800,  "RC",  "Purchase", "⚠ Reverse Charge – overseas digital service"),
        (datetime.date(2025, 10, 22), "Amazon Web Services",     2_400, 0,      2_400,  "RC",  "Purchase", "⚠ Reverse Charge – AWS cloud services"),
        (datetime.date(2025, 11,  5), "Microsoft Azure",         3_100, 0,      3_100,  "RC",  "Purchase", "⚠ Reverse Charge – Azure subscription"),
        (datetime.date(2025, 11, 12), "Salesforce Inc",          1_650, 0,      1_650,  "RC",  "Purchase", "⚠ Reverse Charge – US SaaS"),
        (datetime.date(2025, 11, 18), "Adobe Systems Ltd",         840, 0,        840,  "RC",  "Purchase", "⚠ Reverse Charge – creative software"),
        (datetime.date(2025, 10, 28), "Corporate Golf Day",      1_200, 240,    1_440,  "BLK", "Purchase", "⚠ BLOCKED – client entertainment"),
        (datetime.date(2025, 11, 30), "Christmas Staff Party",   4_800, 960,    5_760,  "BLK", "Purchase", "⚠ BLOCKED – staff entertainment"),
        (datetime.date(2025, 12,  8), "Company Car – BMW 320d", 22_000, 2_200, 24_200,  "BLK", "Purchase", "⚠ 50% BLOCKED – car available for private use"),
        (datetime.date(2025, 10, 20), "BuildRight Ltd",          8_400, 0,      8_400,  "RC",  "Purchase", "⚠ Construction Domestic Reverse Charge"),
        (datetime.date(2025, 11, 25), "SteelFrame Contractors",  5_600, 0,      5_600,  "RC",  "Purchase", "⚠ Construction Domestic Reverse Charge"),
        (datetime.date(2025, 12, 15), "Business Lunch – Client", 420,   0,        420,  "EX",  "Purchase", "⚠ Client entertainment – input VAT blocked"),
        (datetime.date(2025, 10, 10), "Fuel – Company Cars",    1_800, 360,    2_160,  "STD", "Purchase", "Note: Fuel scale charge required"),
        (datetime.date(2025, 11, 20), "Office Supplies",         None,  None,   None,  "",    "Purchase", "⚠ Missing VAT code"),
        (datetime.date(2025, 12,  5), "Amazon Business",         None,  None,   None,  "",    "Purchase", "⚠ Missing VAT code"),
        (datetime.date(2025, 10, 30), "Cleaning Services",       None,  None,   None,  "",    "Purchase", "⚠ Missing VAT code"),
    ]

    rows = list(injected)
    vat_codes = ["STD","ZR","EX"]
    std_net = sum(r[2] or 0 for r in injected if r[5] == "STD")
    target_box6 = VAT_BOX6
    # Add standard-rated sales to hit Box 6
    for i in range(220):
        d = datetime.date(2025, random.randint(10, 12), random.randint(1, 28))
        net = random.randint(2000, 45000)
        std_net += net
        vat = round(net * 0.20, 2)
        code = random.choices(["STD","ZR","EX"], weights=[70, 20, 10])[0]
        vat_ = round(net * 0.20, 2) if code == "STD" else 0
        rows.append((d, fake.company()[:40], net, vat_, net + vat_, code,
                     random.choice(["Sale","Sale","Sale","Purchase"]), ""))

    # Sort by date
    rows.sort(key=lambda r: r[0])

    for ri, row in enumerate(rows[:300], start=4):
        date_, sup, net, vat_, gross, code, typ, note = row
        is_flag = "⚠" in (note or "")
        row_data = [date_, sup, net, vat_, gross, code, typ, note]
        for ci, v in enumerate(row_data, 1):
            c = ws.cell(ri, ci, v)
            c.font = Font(name="Calibri", size=10)
            if ci == 1: c.number_format = DATE_FMT
            if ci in (3, 4, 5) and v is not None:
                c.number_format = GBP0; c.alignment = Alignment(horizontal="right")
            if ci == 6 and code in ("RC","BLK","EX"):
                c.font = Font(name="Calibri", size=10, bold=True,
                    color="C00000" if code in ("BLK","EX") else "1F497D")
            if is_flag:
                c.fill = PatternFill("solid", fgColor="FFF2CC")

    col_w(ws, {"A": 14, "B": 40, "C": 14, "D": 14, "E": 14, "F": 8, "G": 12, "H": 55})
    add_table(ws, f"A3:H{3 + len(rows[:300])}", "VATTransTable")

# ── 11. Bank Transactions ─────────────────────────────────────────────────────

def sheet_bank(wb):
    ws = wb.create_sheet("Bank Transactions")
    ws.sheet_view.showGridLines = False
    freeze(ws, 4, 1)
    ws.merge_cells("A1:F1")
    hdr(ws, 1, 1, f"{COMPANY}  |  Bank Transactions  |  Q4 2025", bg=DARK_BLUE, size=12)

    headers = ["Date","Description","Debit £","Credit £","Balance £","Flag"]
    for ci, h in enumerate(headers, 1):
        hdr(ws, 3, ci, h, bg=MID_BLUE, align="right" if ci > 2 else "left")

    balance = 310_000.0
    rows = []
    base_date = datetime.date(2025, 10, 1)

    # Inject some round-number and duplicate transactions
    rows.append((base_date + datetime.timedelta(2), "Opening Balance",          0, 310_000, 310_000, ""))
    rows.append((base_date + datetime.timedelta(5), "Customer Receipts – Batch 1", 0, 145_000, None, ""))
    rows.append((base_date + datetime.timedelta(7), "BACS – SteelCo Ltd",     85_000, 0, None, ""))
    rows.append((base_date + datetime.timedelta(8), "Round Journal – Suspense",50_000, 0, None, "⚠ ROUND NUMBER"))
    rows.append((base_date + datetime.timedelta(8), "Round Journal – Suspense",50_000, 0, None, "⚠ DUPLICATE AMOUNT"))
    rows.append((base_date + datetime.timedelta(12),"Payroll – October",      215_000, 0, None, ""))
    rows.append((base_date + datetime.timedelta(15),"HMRC VAT Payment",        31_700, 0, None, ""))
    rows.append((base_date + datetime.timedelta(18),"Customer Receipts – Batch 2", 0, 220_000, None, ""))
    rows.append((base_date + datetime.timedelta(22),"Power Systems UK",        48_000, 0, None, ""))
    rows.append((base_date + datetime.timedelta(28),"Barclays Loan Repayment", 35_000, 0, None, ""))
    # Weekend transaction
    weekend = base_date + datetime.timedelta(days=(5 - base_date.weekday()) % 7 + 5)
    rows.append((weekend, "Manual Adjustment – Director",15_000, 0, None, "⚠ WEEKEND POSTING"))

    # Add 239 more transactions
    for i in range(239):
        d = base_date + datetime.timedelta(days=random.randint(0, 91))
        is_receipt = random.random() > 0.55
        amt = random.randint(500, 48_000)
        rows.append((d, fake.company()[:45] + (" Payment" if not is_receipt else " Receipt"),
                     0 if is_receipt else amt, amt if is_receipt else 0, None, ""))

    rows.sort(key=lambda r: r[0])

    # Recalculate balances
    bal = 310_000.0
    for ri, row in enumerate(rows[:250], start=4):
        date_, desc, debit, credit, _, flag = row
        bal = bal - debit + credit
        is_flag = bool(flag)
        row_data = [date_, desc, debit or None, credit or None, bal, flag]
        for ci, v in enumerate(row_data, 1):
            c = ws.cell(ri, ci, v)
            c.font = Font(name="Calibri", size=10)
            if ci == 1: c.number_format = DATE_FMT
            if ci in (3, 4, 5) and v is not None:
                c.number_format = GBP0; c.alignment = Alignment(horizontal="right")
            if ci == 5 and isinstance(v, (int, float)) and v < 0:
                c.font = Font(name="Calibri", size=10, bold=True, color="C00000")
            if is_flag:
                c.fill = PatternFill("solid", fgColor="FFF2CC")

    col_w(ws, {"A": 14, "B": 50, "C": 16, "D": 16, "E": 16, "F": 30})

# ── 12. Bank Reconciliation ───────────────────────────────────────────────────

def sheet_bank_recon(wb):
    ws = wb.create_sheet("Bank Reconciliation")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:D1")
    hdr(ws, 1, 1, f"{COMPANY}  |  Bank Reconciliation  |  31 Dec 2025", bg=DARK_BLUE, size=12)

    # Summary
    summary = [
        ("Bank Statement Balance",         BANK_STMT,  None),
        ("Add: Outstanding Deposits",         18_000,  None),
        ("Less: Unpresented Cheques",         (3_000), None),
        ("ADJUSTED BANK STATEMENT BALANCE", BANK_STMT + 18_000 - 3_000, MID_BLUE),
        ("", None, None),
        ("TB Bank Balance",                   BANK,    None),
        ("", None, None),
        ("RECONCILING DIFFERENCE",            BANK - (BANK_STMT + 18_000 - 3_000), "C00000"),
    ]

    hdr(ws, 3, 1, "Item", bg=MID_BLUE); hdr(ws, 3, 2, "£", bg=MID_BLUE, align="right")
    hdr(ws, 3, 3, "Status", bg=MID_BLUE); hdr(ws, 3, 4, "Notes", bg=MID_BLUE)

    for ri, (label, val_, color) in enumerate(summary, start=4):
        if label == "": ws.row_dimensions[ri].height = 6; continue
        is_diff = label == "RECONCILING DIFFERENCE"
        bg = color if color else WHITE
        fg = WHITE if color and color not in (WHITE,) else ("C00000" if is_diff else "000000")
        ws.cell(ri, 1, label).font = Font(bold=is_diff or bool(color), name="Calibri", size=10, color=fg)
        ws.cell(ri, 1).fill = PatternFill("solid", fgColor=bg)
        c = ws.cell(ri, 2, val_); c.number_format = GBP0
        c.font = Font(bold=is_diff or bool(color), name="Calibri", size=10, color=fg)
        c.fill = PatternFill("solid", fgColor=bg); c.alignment = Alignment(horizontal="right")
        if is_diff:
            ws.cell(ri, 3, "UNRECONCILED").font = Font(bold=True, color="C00000")
            ws.cell(ri, 4, "⚠ 6 unreconciled items below — oldest 97 days").font = Font(bold=True, color="C00000")

    # Unreconciled items
    hdr(ws, 15, 1, "UNRECONCILED ITEMS", bg=AMBER, fg="000000")
    hdr(ws, 16, 1, "Date", bg=MID_BLUE); hdr(ws, 16, 2, "Description", bg=MID_BLUE)
    hdr(ws, 16, 3, "Amount £", bg=MID_BLUE, align="right"); hdr(ws, 16, 4, "Age (Days)", bg=MID_BLUE)
    hdr(ws, 16, 5, "Action Required", bg=MID_BLUE)

    unrec = [
        (datetime.date(2025, 9, 25), "Customer Payment – unallocated", 8_200, 97, "Allocate to debtor account"),
        (datetime.date(2025,10, 12), "Bank Charge – unposted",           420, 80, "Post to bank charges nominal"),
        (datetime.date(2025,10, 30), "BACS Return – unknown supplier",  3_150, 62, "Investigate and re-issue"),
        (datetime.date(2025,11, 15), "Standing Order – unmatched",      1_800, 46, "Identify payee and post"),
        (datetime.date(2025,12,  5), "Direct Debit – uncoded",          1_430, 26, "Code to correct nominal"),
        (datetime.date(2025,12, 20), "Deposit – pending allocation",    1_000, 11, "Allocate to customer account"),
    ]

    for ri, (date_, desc, amt, age, action) in enumerate(unrec, start=17):
        ws.cell(ri, 1, date_).number_format = DATE_FMT
        ws.cell(ri, 2, desc).font = Font(name="Calibri", size=10)
        ws.cell(ri, 3, amt).number_format = GBP0; ws.cell(ri, 3).alignment = Alignment(horizontal="right")
        ws.cell(ri, 4, age)
        ws.cell(ri, 5, action)
        if age > 60:
            for ci in range(1, 6):
                ws.cell(ri, ci).fill = PatternFill("solid", fgColor="FCE4D6")

    col_w(ws, {"A": 14, "B": 45, "C": 16, "D": 14, "E": 45})

# ── 13. Payroll Summary ───────────────────────────────────────────────────────

def sheet_payroll(wb):
    ws = wb.create_sheet("Payroll Summary")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:G1")
    hdr(ws, 1, 1, f"{COMPANY}  |  Payroll Summary  |  December 2025", bg=DARK_BLUE, size=12)

    headers = ["Department","Headcount","Gross Pay £","Employer NIC £","Pension £","Total Cost £","TB Posted"]
    for ci, h in enumerate(headers, 1):
        hdr(ws, 3, ci, h, bg=MID_BLUE, align="right" if ci > 2 else "left")

    depts = [
        ("Operations",   28, 64_000),
        ("Production",   22, 52_000),
        ("Sales",        10, 38_000),
        ("Admin & Finance", 8, 30_000),
        ("Management",    7, 31_000),
    ]

    for ri, (dept, hc, gross) in enumerate(depts, start=4):
        nic  = round(gross * 0.138, 0)
        pen  = round(gross * 0.04, 0)
        tot  = gross + nic + pen
        # Inject: payroll NOT posted to TB this month
        posted = "⚠ NOT POSTED" if dept == "Production" else "✓ Posted"
        row_data = [dept, hc, gross, nic, pen, tot, posted]
        for ci, v in enumerate(row_data, 1):
            c = ws.cell(ri, ci, v)
            c.font = Font(name="Calibri", size=10, color="C00000" if posted == "⚠ NOT POSTED" else "000000")
            if ci in (3,4,5,6): c.number_format = GBP0; c.alignment = Alignment(horizontal="right")
            if posted == "⚠ NOT POSTED":
                c.fill = PatternFill("solid", fgColor="FCE4D6")

    tot_row = 9
    hdr(ws, tot_row, 1, "TOTAL", bg=MID_BLUE)
    ws.cell(tot_row, 2, EMPLOYEES)
    ws.cell(tot_row, 3, MONTHLY_PAY).number_format = GBP0
    ws.cell(tot_row, 4, round(MONTHLY_PAY * 0.138)).number_format = GBP0
    ws.cell(tot_row, 5, round(MONTHLY_PAY * 0.04)).number_format = GBP0
    ws.cell(tot_row, 6, round(MONTHLY_PAY * 1.178)).number_format = GBP0

    ws.cell(12, 1, "⚠ WARNING: Production payroll journal has NOT been posted to the trial balance this month.").font = Font(bold=True, color="C00000")
    ws.cell(12, 1).fill = PatternFill("solid", fgColor="FCE4D6")
    ws.merge_cells("A12:G12")

    col_w(ws, {"A": 22, "B": 12, "C": 16, "D": 16, "E": 12, "F": 16, "G": 18})

# ── 14. Fixed Asset Register ──────────────────────────────────────────────────

def sheet_fixed_assets(wb):
    ws = wb.create_sheet("Fixed Asset Register")
    ws.sheet_view.showGridLines = False
    freeze(ws, 4, 1)
    ws.merge_cells("A1:J1")
    hdr(ws, 1, 1, f"{COMPANY}  |  Fixed Asset Register  |  {YEAR_END}", bg=DARK_BLUE, size=12)

    headers = ["Asset Code","Asset Description","Category","Acquisition Date","Cost £",
               "Useful Life (Yrs)","Annual Depn £","Accum Depn £","NBV £","Flag"]
    for ci, h in enumerate(headers, 1):
        hdr(ws, 3, ci, h, bg=MID_BLUE, align="right" if ci > 4 else "left")

    # Categories
    assets = []
    # Plant & Machinery
    pm_items = [
        ("FA001","CNC Milling Machine – Mazak","Plant & Machinery", datetime.date(2018, 3,15), 285_000, 15),
        ("FA002","Press Brake – Trumpf","Plant & Machinery",        datetime.date(2019, 6,20), 180_000, 12),
        ("FA003","Welding Robots – ABB x4","Plant & Machinery",     datetime.date(2020, 1,10), 420_000, 15),
        ("FA004","Forklift Trucks x3","Plant & Machinery",          datetime.date(2021, 4, 5), 145_000, 10),
        ("FA005","Conveyor System","Plant & Machinery",              datetime.date(2017, 8,12), 320_000, 20),
        # Zero depreciation injected
        ("FA006","Laser Cutter – Trumpf 5000","Plant & Machinery",  datetime.date(2023, 9,15), 380_000, 15),
        ("FA007","Industrial Compressor Array","Plant & Machinery",  datetime.date(2023,11, 1), 250_000, 12),
        ("FA008","Heat Treatment Furnace","Plant & Machinery",       datetime.date(2024, 2,20), 120_000, 20),
    ]
    mv_items = [
        ("FA020","Ford Transit Van – NG22 ABC","Motor Vehicles",  datetime.date(2022, 5,10), 38_000, 5),
        ("FA021","VW Caddy Van – YJ23 DEF","Motor Vehicles",      datetime.date(2023, 3,15), 32_000, 5),
        ("FA022","BMW 5 Series – SA23 GHI","Motor Vehicles",      datetime.date(2023, 6,20), 52_000, 4),
        ("FA023","Toyota Hilux – LK24 JKL","Motor Vehicles",      datetime.date(2024, 1,15), 45_000, 5),
    ]
    it_items = [
        ("FA030","ERP System – Sage 200","IT Equipment",          datetime.date(2020, 9, 1),  85_000, 7),
        ("FA031","Servers – Dell PowerEdge x4","IT Equipment",    datetime.date(2021, 3,15),  48_000, 5),
        ("FA032","Laptops x45 – HP Elite","IT Equipment",         datetime.date(2023, 9, 1),  47_000, 3),
    ]

    ZERO_DEPN_ASSETS = {"FA006","FA007","FA008"}  # £750k with zero depreciation

    for code, desc, cat, acq_date, cost, life in [*pm_items, *mv_items, *it_items]:
        years_held = max(0, (datetime.date(2025,12,31) - acq_date).days / 365)
        if code in ZERO_DEPN_ASSETS:
            annual_depn = 0
            accum_depn  = 0
            flag = "⚠ ZERO DEPRECIATION – review required"
        else:
            annual_depn = round(cost / life)
            accum_depn  = round(min(annual_depn * years_held, cost))
            flag = ""
        nbv = cost - accum_depn
        assets.append((code, desc, cat, acq_date, cost, life, annual_depn, accum_depn, nbv, flag))

    for ri, (code, desc, cat, acq, cost, life, adepn, acdepn, nbv, flag) in enumerate(assets, start=4):
        row_data = [code, desc, cat, acq, cost, life, adepn, acdepn, nbv, flag]
        is_flag = bool(flag)
        for ci, v in enumerate(row_data, 1):
            c = ws.cell(ri, ci, v)
            c.font = Font(name="Calibri", size=10, color="C00000" if is_flag and ci in (7,8,10) else "000000")
            if ci == 4: c.number_format = DATE_FMT
            if ci in (5,7,8,9): c.number_format = GBP0; c.alignment = Alignment(horizontal="right")
            if ci == 6: c.alignment = Alignment(horizontal="right")
            if is_flag: c.fill = PatternFill("solid", fgColor="FCE4D6")

    # Totals
    tot = 4 + len(assets)
    hdr(ws, tot, 1, "TOTALS", bg=DARK_BLUE)
    for ci, val_ in enumerate([
            sum(a[4] for a in assets), "", sum(a[6] for a in assets),
            sum(a[7] for a in assets), sum(a[8] for a in assets)
    ], start=5):
        if val_ == "": continue
        c = ws.cell(tot, ci, val_); c.number_format = GBP0
        c.font = Font(bold=True, name="Calibri", size=10, color=WHITE)
        c.fill = PatternFill("solid", fgColor=DARK_BLUE); c.alignment = Alignment(horizontal="right")

    # Warning
    ws.cell(tot+2, 1, "⚠ FA006, FA007, FA008 — Cost £750,000 — NO DEPRECIATION CHARGED. Review urgently.").font = Font(bold=True, color="C00000")
    ws.merge_cells(f"A{tot+2}:J{tot+2}")

    col_w(ws, {"A": 8, "B": 40, "C": 20, "D": 16, "E": 14, "F": 14, "G": 16, "H": 16, "I": 14, "J": 45})
    add_table(ws, f"A3:J{tot-1}", "FARegTable")

# ── 15. Cashflow Forecast ─────────────────────────────────────────────────────

def sheet_cashflow(wb):
    ws = wb.create_sheet("Cashflow Forecast")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:E1")
    hdr(ws, 1, 1, f"{COMPANY}  |  13-Week Cashflow Forecast  |  Q1 2026", bg=DARK_BLUE, size=12)

    headers = ["Week","Period","Opening Cash £","Receipts £","Payments £","Closing Cash £","Status"]
    for ci, h in enumerate(headers, 1):
        hdr(ws, 3, ci, h, bg=MID_BLUE, align="right" if ci > 2 else "left")

    opening = BANK
    start = datetime.date(2026, 1, 5)

    rows = []
    for w in range(1, 14):
        period = f"w/e {start + datetime.timedelta(weeks=w-1)}"
        if w == 8:
            receipts = 28_000
            payments = 125_000
        else:
            receipts = random.randint(60_000, 180_000)
            payments = random.randint(55_000, 170_000)
        closing = opening + receipts - payments
        status = "🔴 NEGATIVE" if closing < 0 else ("🟡 LOW" if closing < 50_000 else "🟢 OK")
        rows.append((w, period, opening, receipts, payments, closing, status))
        opening = closing

    for ri, (w, period, op, rec, pay, cl, status) in enumerate(rows, start=4):
        is_neg = cl < 0
        row_data = [w, period, op, rec, pay, cl, status]
        for ci, v in enumerate(row_data, 1):
            c = ws.cell(ri, ci, v)
            c.font = Font(name="Calibri", size=10, color="C00000" if is_neg and ci in (3,4,5,6) else "000000")
            if ci in (3,4,5,6): c.number_format = GBP0; c.alignment = Alignment(horizontal="right")
            if is_neg: c.fill = PatternFill("solid", fgColor="FCE4D6")
        if is_neg:
            ws.cell(ri, 6).font = Font(bold=True, color="C00000", name="Calibri", size=10)

    ws.cell(18, 1, "⚠ Week 8: Negative cash balance projected. Action required to manage collections and payment timing.").font = Font(bold=True, color="C00000")

    col_w(ws, {"A": 8, "B": 22, "C": 18, "D": 16, "E": 16, "F": 18, "G": 16})

# ── 16. Cross File Reconciliation ─────────────────────────────────────────────

def sheet_cross_recon(wb):
    ws = wb.create_sheet("Cross File Reconciliation")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    hdr(ws, 1, 1, f"{COMPANY}  |  Cross-File Reconciliation Summary  |  {YEAR_END}", bg=DARK_BLUE, size=12)

    headers = ["Check","Source A","Source B","Value A £","Value B £","Difference £","Result"]
    for ci, h in enumerate(headers, 1):
        hdr(ws, 3, ci, h, bg=MID_BLUE, align="right" if ci > 3 else "left")

    checks = [
        ("AR Ledger ↔ Debtors Control",    "AR Aging",       "Trial Balance",      AR_AGING_TOTAL,  TRADE_DEBTORS, AR_AGING_TOTAL - TRADE_DEBTORS, "FAIL"),
        ("AP Ledger ↔ Creditors Control",  "AP Aging",       "Trial Balance",      TRADE_CRED,      TRADE_CRED,    0,                              "PASS"),
        ("VAT Return ↔ VAT Control",        "VAT Return",     "Trial Balance",      VAT_BOX1-VAT_BOX4, VAT_CTRL,  (VAT_BOX1-VAT_BOX4) - VAT_CTRL, "FAIL"),
        ("Bank ↔ Trial Balance",            "Bank Statement", "Trial Balance",      BANK_STMT,       BANK,          BANK_STMT - BANK,               "FAIL"),
        ("Balance Sheet Equation",          "Total Assets",   "Liabilities+Equity", TOTAL_ASSETS,    TOTAL_LIAB+TOTAL_EQUITY, TOTAL_ASSETS-(TOTAL_LIAB+TOTAL_EQUITY), "PASS"),
        ("P&L ↔ Equity Movement",           "Net Profit",     "Retained Earnings",  NET_PROFIT,      NET_PROFIT,   0,                              "PASS"),
        ("Payroll ↔ TB Payroll Account",    "Payroll Summary","Trial Balance",      MONTHLY_PAY,     0,             MONTHLY_PAY,                    "FAIL"),
    ]

    for ri, (check, srcA, srcB, valA, valB, diff, result) in enumerate(checks, start=4):
        is_fail = result == "FAIL"
        row_data = [check, srcA, srcB, valA, valB, diff, result]
        for ci, v in enumerate(row_data, 1):
            c = ws.cell(ri, ci, v)
            c.font = Font(name="Calibri", size=10, bold=(ci == 7))
            if ci in (4,5,6): c.number_format = GBP0; c.alignment = Alignment(horizontal="right")
            if ci == 7:
                c.font = Font(bold=True, name="Calibri", size=10,
                    color="C00000" if is_fail else "375623")
                c.fill = PatternFill("solid", fgColor="FCE4D6" if is_fail else "E2EFDA")
            if is_fail and ci != 7:
                c.fill = PatternFill("solid", fgColor="FFF2CC")

    col_w(ws, {"A": 35, "B": 20, "C": 20, "D": 18, "E": 18, "F": 16, "G": 10})

# ── 17. Expected Findings ─────────────────────────────────────────────────────

def sheet_findings(wb):
    ws = wb.create_sheet("Expected Findings")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:G1")
    hdr(ws, 1, 1, f"{COMPANY}  |  Expected ClosePilot Findings  |  {YEAR_END}", bg=DARK_BLUE, size=12)

    headers = ["Finding ID","Finding","Severity","Category","Exposure £","Rule ID","Description"]
    for ci, h in enumerate(headers, 1):
        hdr(ws, 3, ci, h, bg=MID_BLUE, align="right" if ci == 5 else "left")

    findings = [
        ("F001","AR Ledger vs Debtors Control Mismatch","High","AR",       45_000, "AR_CTRL_01","AR Aging total £895k vs TB debtors control £850k — £45k unreconciled"),
        ("F002","VAT Return vs VAT Control Account",     "High","VAT",       8_400, "VAT_051",   "VAT return net £31,700 vs TB VAT control £42,000 — £10,300 difference"),
        ("F003","Bank Statement vs TB Balance",          "High","Controls",  15_000, "DI_004",   "Bank statement £325k vs TB bank £340k — 6 unreconciled items, oldest 97 days"),
        ("F004","Production Payroll Not Posted",         "High","Month End", 52_000, "CR_008",   "December production payroll £52k not posted to trial balance"),
        ("F005","Fixed Assets – Zero Depreciation",     "High","FS",        750_000,"FS_005",   "3 assets totalling £750k acquired 2023-2024 with no depreciation charged"),
        ("F006","Dunlop Retail – 132 Days Overdue",     "Critical","AR",   245_000, "AR_003",   "£245k outstanding 132 days — provision required. Exceeds credit limit."),
        ("F007","Construction Reverse Charge Missing",  "High","VAT",        14_000, "VAT_009",  "BuildRight Ltd and SteelFrame: construction services not reverse charged"),
        ("F008","Blocked VAT on Entertainment",         "Medium","VAT",       1_200, "VAT_004",  "Corporate golf day £1,200 and Christmas party £960 input VAT — blocked"),
        ("F009","Company Car VAT – 50% Block Required", "Medium","VAT",       2_200, "VAT_012",  "BMW 320d purchase £22,000 — 50% input VAT block required (£2,200)"),
        ("F010","Digital Services Reverse Charge",      "High","VAT",         9_790, "VAT_010",  "Google/AWS/Azure/Salesforce/Adobe — £9,790 net — RC accounting required"),
        ("F011","Duplicate Supplier – ABC Services",    "Medium","AP",        12_400, "AP_001",  "ABC Services Ltd invoice reference duplicated — possible duplicate payment"),
        ("F012","Personal Payee – Mr John Smith",       "Medium","Controls",  4_950, "CF_009",  "Payment to individual — IR35/PAYE review required"),
        ("F013","Supplier Over 90 Days – Meridian",     "Medium","AP",         8_200, "AP_004",  "Meridian Tools Ltd 185 days overdue — consider write-back"),
        ("F014","Interest Cover at Risk",               "High","Cashflow",   120_000,"ST_028",  "Interest charge £120k vs EBIT £60k — interest cover 0.5x (below 2x covenant)"),
        ("F015","Customer Concentration – Dunlop",      "High","AR",         245_000,"AR_002",  "Dunlop Retail represents 27% of total AR — concentration risk"),
        ("F016","Negative Cash – Week 8 Forecast",     "Medium","Cashflow",   50_000,"CF_001",  "13-week cashflow shows negative position in Week 8 — action required"),
        ("F017","Fuel Scale Charge Missing",            "Medium","VAT",          360, "VAT_005", "Company car fuel VAT claimed — fuel scale charge not detected"),
        ("F018","Round-Number Bank Posting",            "Low","Controls",      50_000,"CF_002",  "£50,000 round-number journal posted to suspense — appears twice"),
        ("F019","Weekend Bank Posting – Director",      "Medium","Controls",   15_000,"CF_001",  "Manual adjustment £15k posted Saturday — Director authorisation?"),
        ("F020","Missing VAT Codes – 3 Transactions",  "Medium","VAT",          None, "VAT_001", "Office Supplies, Amazon Business, Cleaning Services have blank VAT codes"),
    ]

    sev_colors = {"Critical":"C00000","High":"FF0000","Medium":"FF8C00","Low":"FFC000"}

    for ri, (fid, finding, sev, cat, exp, rule, desc) in enumerate(findings, start=4):
        row_data = [fid, finding, sev, cat, exp, rule, desc]
        for ci, v in enumerate(row_data, 1):
            c = ws.cell(ri, ci, v)
            c.font = Font(name="Calibri", size=10)
            if ci == 3:
                c.font = Font(bold=True, name="Calibri", size=10, color=sev_colors.get(sev,"000000"))
            if ci == 5 and v:
                c.number_format = GBP0; c.alignment = Alignment(horizontal="right")
        ws.cell(ri, 1).fill = PatternFill("solid", fgColor=LIGHT_GREY)

    col_w(ws, {"A": 8, "B": 50, "C": 10, "D": 12, "E": 16, "F": 12, "G": 65})
    add_table(ws, f"A3:G{3+len(findings)}", "FindingsTable")

# ── 18. Review Pack Summary ───────────────────────────────────────────────────

def sheet_summary(wb):
    ws = wb.create_sheet("Review Pack Summary")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:D1")
    hdr(ws, 1, 1, f"CLOSEPILOT AUDIT READINESS REPORT  |  {COMPANY}", bg=DARK_BLUE, size=14)
    hdr(ws, 2, 1, f"Prepared by: ClosePilot Assurance Platform  |  {YEAR_END}", bg=MID_BLUE)
    ws.merge_cells("A2:D2")

    # Key metrics
    metrics = [
        ("Audit Readiness Score",     "79%",  MID_BLUE),
        ("Close Readiness",           "83%",  MID_BLUE),
        ("Confidence Score",          "88%",  MID_BLUE),
        ("Finance Health Score",      "72%",  MID_BLUE),
    ]

    counts = [
        ("Critical Findings",   0,  GREEN_FILL),
        ("High Findings",       7,  "FF0000"),
        ("Medium Findings",    13,  AMBER),
        ("Total Findings",     20,  DARK_BLUE),
    ]

    exposure = [
        ("AR Risk",    450_000),
        ("VAT Risk",     8_400),
        ("Bank Risk",   15_000),
        ("Payroll Risk",52_000),
        ("Asset Risk",  750_000),  # zero depreciation
        ("Other",       44_700),
    ]

    # Left column – scores
    hdr(ws, 5, 1, "SCORES", bg=DARK_BLUE)
    ws.merge_cells("A5:B5")
    for ri, (label, val_, color) in enumerate(metrics, start=6):
        ws.cell(ri, 1, label).font = Font(bold=True, name="Calibri", size=12)
        ws.cell(ri, 1).fill = PatternFill("solid", fgColor=LIGHT_GREY)
        c = ws.cell(ri, 2, val_)
        c.font = Font(bold=True, name="Calibri", size=20, color=WHITE)
        c.fill = PatternFill("solid", fgColor=color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[ri].height = 30

    # Right column – counts
    hdr(ws, 5, 3, "FINDINGS", bg=DARK_BLUE)
    ws.merge_cells("C5:D5")
    for ri, (label, count, color) in enumerate(counts, start=6):
        ws.cell(ri, 3, label).font = Font(bold=True, name="Calibri", size=12)
        ws.cell(ri, 3).fill = PatternFill("solid", fgColor=LIGHT_GREY)
        c = ws.cell(ri, 4, count)
        c.font = Font(bold=True, name="Calibri", size=20, color=WHITE)
        c.fill = PatternFill("solid", fgColor=color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[ri].height = 30

    # Financial exposure
    hdr(ws, 12, 1, "FINANCIAL EXPOSURE", bg=DARK_BLUE)
    ws.merge_cells("A12:D12")
    total_exp = sum(v for _, v in exposure)
    for ri, (label, amt) in enumerate(exposure, start=13):
        ws.cell(ri, 1, label).font = Font(name="Calibri", size=11, bold=True)
        c = ws.cell(ri, 2, amt); c.number_format = GBP0
        c.font = Font(name="Calibri", size=11); c.alignment = Alignment(horizontal="right")
    hdr(ws, 13 + len(exposure), 1, "TOTAL FINANCIAL EXPOSURE", bg=MID_BLUE)
    c = ws.cell(13 + len(exposure), 2, total_exp); c.number_format = GBP0
    c.font = Font(bold=True, name="Calibri", size=12, color=WHITE)
    c.fill = PatternFill("solid", fgColor=MID_BLUE); c.alignment = Alignment(horizontal="right")

    # Key recommendations
    hdr(ws, 21, 1, "PRIORITY ACTIONS", bg=DARK_BLUE)
    ws.merge_cells("A21:D21")
    actions = [
        "1. Investigate AR vs Debtors Control mismatch — £45,000 unreconciled",
        "2. Post production payroll journal — £52,000 missing from TB",
        "3. Review and charge depreciation on FA006/FA007/FA008 — £750,000 assets",
        "4. Complete VAT reverse charge accounting before return submission",
        "5. Issue formal demand for Dunlop Retail — £245,000 at 132 days",
        "6. Resolve 6 bank reconciliation items — oldest 97 days",
        "7. Review interest cover covenant — currently 0.5x vs 2.0x minimum",
    ]
    for ri, action in enumerate(actions, start=22):
        ws.cell(ri, 1, action).font = Font(name="Calibri", size=10, bold=(ri == 22))
        ws.merge_cells(f"A{ri}:D{ri}")
        if ri % 2 == 0:
            ws.cell(ri, 1).fill = PatternFill("solid", fgColor=LIGHT_GREY)

    ws.cell(30, 1, f"Generated by ClosePilot Assurance Platform  |  {datetime.date.today().strftime('%d %B %Y')}  |  Confidential").font = Font(name="Calibri", size=9, italic=True, color="7F7F7F")
    ws.merge_cells("A30:D30")

    col_w(ws, {"A": 38, "B": 18, "C": 28, "D": 16})

# ── Assemble and save ─────────────────────────────────────────────────────────

def main():
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    print("Generating worksheets...")
    sheet_company_profile(wb)   ; print("  1/18 Company Profile")
    sheet_trial_balance(wb)     ; print("  2/18 Trial Balance (250+ accounts)")
    sheet_balance_sheet(wb)     ; print("  3/18 Balance Sheet")
    sheet_pnl(wb)               ; print("  4/18 Profit & Loss")
    sheet_12month(wb)           ; print("  5/18 12 Month P&L History")
    sheet_budget(wb)            ; print("  6/18 Budget vs Actual")
    sheet_ar_aging(wb)          ; print("  7/18 AR Aging (50 customers, £895k, £45k mismatch)")
    sheet_ap_aging(wb)          ; print("  8/18 AP Aging (50 suppliers, duplicate, personal payee)")
    sheet_vat_return(wb)        ; print("  9/18 VAT Return (Q4 2025, £8.4k mismatch)")
    sheet_vat_transactions(wb)  ; print(" 10/18 VAT Transactions (300 rows)")
    sheet_bank(wb)              ; print(" 11/18 Bank Transactions (250 rows)")
    sheet_bank_recon(wb)        ; print(" 12/18 Bank Reconciliation (£15k difference)")
    sheet_payroll(wb)           ; print(" 13/18 Payroll Summary (75 employees)")
    sheet_fixed_assets(wb)      ; print(" 14/18 Fixed Asset Register (£750k zero depreciation)")
    sheet_cashflow(wb)          ; print(" 15/18 Cashflow Forecast (13 weeks, negative Week 8)")
    sheet_cross_recon(wb)       ; print(" 16/18 Cross File Reconciliation")
    sheet_findings(wb)          ; print(" 17/18 Expected Findings (20 findings)")
    sheet_summary(wb)           ; print(" 18/18 Review Pack Summary")

    output = "ClosePilot_Enterprise_Demo_Pack.xlsx"
    wb.save(output)
    print(f"\n✓ Saved: {output}")
    print(f"  Sheets: {len(wb.worksheets)}")

    # Count rows
    total_rows = sum(ws.max_row for ws in wb.worksheets)
    print(f"  Total rows: {total_rows:,}")
    print("\nDeliberate issues injected:")
    print(f"  AR mismatch:    £{AR_MISMATCH:,} (aging £{AR_AGING_TOTAL:,} vs TB £{TRADE_DEBTORS:,})")
    print(f"  VAT mismatch:   £{abs((VAT_BOX1-VAT_BOX4) - VAT_CTRL):,}")
    print(f"  Bank recon:     £{BANK_UNRECON:,} difference ({BANK_STMT:,} vs {BANK:,})")
    print(f"  Zero depreciation: £750,000 on 3 assets")
    print(f"  Payroll not posted: £{MONTHLY_PAY:,}")
    print(f"  Dunlop Retail: £245,000 at 132 days overdue")
    print(f"  Personal payee: Mr John Smith £4,950")
    print(f"  Duplicate invoice: ABC Services")
    print(f"  Reverse charge missing: 7 transactions")


if __name__ == "__main__":
    main()
